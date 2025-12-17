"""
Database Comparison Tool
Compares data from two database sources and generates comparison reports
"""

import pandas as pd
import numpy as np
import logging
from logging.handlers import RotatingFileHandler
import tracemalloc
from datetime import datetime
import os
from typing import List, Dict, Optional, Union
import sqlalchemy
from sqlalchemy import create_engine
from sqlalchemy.engine import Engine


version = "1.0.0"

# Global configuration for pandas display
pd.set_option('display.max_rows', 500)
pd.set_option('display.max_columns', 500)
pd.set_option('display.width', 1000)

# Memory monitoring
tracemalloc.start()


class DataSource:
    """
    Represents a database data source with connection and query configuration
    """

    def __init__(self,
                 name: str,
                 db_type: str,
                 host: str,
                 port: int,
                 database: str,
                 username: str,
                 password: str,
                 sql_query: str,
                 **kwargs):
        """
        Initialize a data source

        Args:
            name: Name of the data source (e.g., "source1", "source2")
            db_type: Database type (mysql, postgresql, oracle, mssql, sqlite)
            host: Database host
            port: Database port
            database: Database name
            username: Database username
            password: Database password
            sql_query: SQL query to fetch data
            **kwargs: Additional connection parameters
        """
        self.name = name
        self.db_type = db_type.lower()
        self.host = host
        self.port = port
        self.database = database
        self.username = username
        self.password = password
        self.sql_query = sql_query
        self.connection_params = kwargs
        self.engine: Optional[Engine] = None

    def get_connection_string(self) -> str:
        """
        Generate SQLAlchemy connection string based on database type

        Returns:
            Connection string for SQLAlchemy
        """
        if self.db_type == 'mysql':
            # MySQL connection string
            # Requires: pip install pymysql
            base_url = f"mysql+pymysql://{self.username}:{self.password}@{self.host}:{self.port}/{self.database}"

            # Add charset and other connection parameters to URL
            params = []
            if 'charset' in self.connection_params:
                params.append(f"charset={self.connection_params['charset']}")

            if params:
                base_url += "?" + "&".join(params)

            return base_url

        elif self.db_type == 'postgresql' or self.db_type == 'postgres':
            # PostgreSQL connection string
            # Requires: pip install psycopg2-binary
            return f"postgresql+psycopg2://{self.username}:{self.password}@{self.host}:{self.port}/{self.database}"

        elif self.db_type == 'oracle':
            # Oracle connection string
            # Requires: pip install cx_oracle
            return f"oracle+cx_oracle://{self.username}:{self.password}@{self.host}:{self.port}/{self.database}"

        elif self.db_type == 'mssql' or self.db_type == 'sqlserver':
            # Microsoft SQL Server connection string
            # Requires: pip install pyodbc
            driver = self.connection_params.get('driver', 'ODBC Driver 17 for SQL Server')
            return f"mssql+pyodbc://{self.username}:{self.password}@{self.host}:{self.port}/{self.database}?driver={driver}"

        elif self.db_type == 'sqlite':
            # SQLite connection string (file-based)
            return f"sqlite:///{self.database}"

        else:
            raise ValueError(f"Unsupported database type: {self.db_type}")

    def connect(self) -> Engine:
        """
        Create database engine connection

        Returns:
            SQLAlchemy Engine object
        """
        if self.engine is None:
            connection_string = self.get_connection_string()

            # For MySQL, charset is already in the URL, so remove it from engine params
            engine_params = self.connection_params.copy()
            if self.db_type == 'mysql' and 'charset' in engine_params:
                engine_params.pop('charset')

            self.engine = create_engine(connection_string, **engine_params)
        return self.engine

    def fetch_data(self) -> pd.DataFrame:
        """
        Execute SQL query and return DataFrame

        Returns:
            DataFrame containing query results
        """
        engine = self.connect()
        df = pd.read_sql_query(self.sql_query, engine)
        return df

    def close(self):
        """Close database connection"""
        if self.engine:
            self.engine.dispose()
            self.engine = None


class DBCompare:
    """
    Database comparison class for comparing data from two database sources
    """

    def __init__(self,
                 data_source1: DataSource,
                 data_source2: DataSource,
                 join_columns: Union[str, List[str]],
                 comparing_columns: Optional[List[str]] = None,
                 column_mapping: Optional[Dict[str, str]] = None,
                 tolerance: Optional[Dict[str, float]] = None,
                 abs_tol: float = 0,
                 rel_tol: float = 0,
                 show_join_columns_both_sides: bool = False,
                 show_transformed_columns: bool = False,
                 output_dir: str = "output",
                 log_file: str = "db_compare.log"):
        """
        Initialize database comparison

        Args:
            data_source1: First data source (DataSource object)
            data_source2: Second data source (DataSource object)
            join_columns: Column(s) to join on (composite key support)
            comparing_columns: Specific columns to compare (None = all columns)
            column_mapping: Dictionary to map column names from source1 to source2 (e.g., {'email1': 'email'})
            tolerance: Dictionary of column-specific tolerances for numeric comparison
            abs_tol: Absolute tolerance for all numeric columns
            rel_tol: Relative tolerance for all numeric columns
            show_join_columns_both_sides: Show join columns on both sides in Excel (default: False, shows once)
            show_transformed_columns: Show transformed column names in Excel (default: False, shows original)
            output_dir: Directory for output files
            log_file: Log file path
        """
        self.data_source1 = data_source1
        self.data_source2 = data_source2

        # Normalize join_columns to list
        if isinstance(join_columns, str):
            self.join_columns = [join_columns]
        else:
            self.join_columns = join_columns

        self.comparing_columns = comparing_columns
        self.column_mapping = column_mapping or {}
        self.tolerance = tolerance or {}
        self.abs_tol = abs_tol
        self.rel_tol = rel_tol
        self.show_join_columns_both_sides = show_join_columns_both_sides
        self.show_transformed_columns = show_transformed_columns
        self.output_dir = output_dir
        self.log_file = log_file

        # Create output directory if it doesn't exist
        os.makedirs(output_dir, exist_ok=True)

        # Setup logging
        self.log = self._create_log()

        # Suffixes for merged columns
        self.suffix1 = "_source1"
        self.suffix2 = "_source2"

        # Results storage
        self.df1: Optional[pd.DataFrame] = None
        self.df2: Optional[pd.DataFrame] = None
        self.df_merged: Optional[pd.DataFrame] = None
        self.comparison_summary: Dict = {}

        # Store original column names before mapping (for Excel display)
        self.original_columns_source1: Optional[List[str]] = None
        self.original_columns_source2: Optional[List[str]] = None

        # Store original join column names before transformation
        # Maps: transformed_name -> original_name for each source
        self.original_join_cols_source1: Dict[str, str] = {}
        self.original_join_cols_source2: Dict[str, str] = {}

    def _create_log(self) -> logging.Logger:
        """
        Create rotating log handler

        Returns:
            Logger object
        """
        logger = logging.getLogger("db_compare")
        logger.setLevel(logging.INFO)

        # Clear existing handlers
        logger.handlers.clear()

        # Formatter
        formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')

        # Console handler
        ch = logging.StreamHandler()
        ch.setLevel(logging.INFO)
        ch.setFormatter(formatter)

        # File handler (rotating)
        fh = RotatingFileHandler(self.log_file, maxBytes=1024*1024*10, backupCount=5)
        fh.setFormatter(formatter)

        logger.addHandler(ch)
        logger.addHandler(fh)

        return logger

    def fetch_data(self):
        """Fetch data from both sources"""
        self.log.info(f"Fetching data from {self.data_source1.name}...")
        self.df1 = self.data_source1.fetch_data()
        self.log.info(f"Fetched {len(self.df1)} rows from {self.data_source1.name}")
        self.log.info(f"Columns before mapping: {list(self.df1.columns)}")

        # Store original columns before mapping
        self.original_columns_source1 = list(self.df1.columns)

        # Apply column mapping if specified
        if self.column_mapping:
            self.log.info(f"Applying column mapping: {self.column_mapping}")
            self.df1 = self.df1.rename(columns=self.column_mapping)
            self.log.info(f"Columns after mapping: {list(self.df1.columns)}")

        self.log.info(f"Fetching data from {self.data_source2.name}...")
        self.df2 = self.data_source2.fetch_data()
        self.log.info(f"Fetched {len(self.df2)} rows from {self.data_source2.name}")
        self.log.info(f"Columns: {list(self.df2.columns)}")

        # Store original columns from source2
        self.original_columns_source2 = list(self.df2.columns)

        # Apply join column transformations
        self._apply_join_column_transforms()

        # Validate join columns exist
        for col in self.join_columns:
            if col not in self.df1.columns:
                raise ValueError(f"Join column '{col}' not found in {self.data_source1.name}")
            if col not in self.df2.columns:
                raise ValueError(f"Join column '{col}' not found in {self.data_source2.name}")

    def _apply_join_column_transforms(self):
        """
        Apply transformation functions to join columns if specified
        Handles both simple string list and complex dictionary format
        """
        from transform_functions import apply_transform

        # If join_columns is a simple list of strings, no transformations needed
        if not self.join_columns:
            return

        # Check if join_columns has transformation specs (list of dicts)
        if isinstance(self.join_columns, list) and len(self.join_columns) > 0:
            if isinstance(self.join_columns[0], dict):
                # Process complex join column specifications
                processed_join_cols = []

                for join_spec in self.join_columns:
                    target_col = join_spec.get('column')
                    source1_col = join_spec.get('source1_column', target_col)
                    source2_col = join_spec.get('source2_column', target_col)
                    source1_transform = join_spec.get('source1_transform')
                    source2_transform = join_spec.get('source2_transform')

                    # Store original column names for Excel display
                    self.original_join_cols_source1[target_col] = source1_col
                    self.original_join_cols_source2[target_col] = source2_col

                    # Apply transformation to source1 if specified
                    if source1_transform and source1_col in self.df1.columns:
                        self.log.info(f"Applying transform '{source1_transform}' to source1 column '{source1_col}'")
                        # Create transformed column (don't drop original)
                        self.df1[target_col] = apply_transform(self.df1[source1_col], source1_transform)
                        # Keep original column with original name
                    elif source1_col != target_col and source1_col in self.df1.columns:
                        # Just rename if no transform
                        self.df1 = self.df1.rename(columns={source1_col: target_col})

                    # Apply transformation to source2 if specified
                    if source2_transform and source2_col in self.df2.columns:
                        self.log.info(f"Applying transform '{source2_transform}' to source2 column '{source2_col}'")
                        # Create transformed column (don't drop original)
                        self.df2[target_col] = apply_transform(self.df2[source2_col], source2_transform)
                        # Keep original column with original name
                    elif source2_col != target_col and source2_col in self.df2.columns:
                        # Just rename if no transform
                        self.df2 = self.df2.rename(columns={source2_col: target_col})

                    processed_join_cols.append(target_col)

                # Update join_columns to the processed column names
                self.join_columns = processed_join_cols
                self.log.info(f"Processed join columns: {self.join_columns}")

        # Filter columns if comparing_columns specified
        if self.comparing_columns:
            # Ensure join columns are included
            cols_to_keep = list(set(self.join_columns + self.comparing_columns))

            # Also keep original join columns if they exist (for show_transformed_columns feature)
            for join_col in self.join_columns:
                orig1 = self.original_join_cols_source1.get(join_col)
                orig2 = self.original_join_cols_source2.get(join_col)
                if orig1 and orig1 != join_col:
                    cols_to_keep.append(orig1)
                if orig2 and orig2 != join_col:
                    cols_to_keep.append(orig2)

            # Remove duplicates
            cols_to_keep = list(set(cols_to_keep))

            # Filter columns that exist
            df1_cols = [c for c in cols_to_keep if c in self.df1.columns]
            df2_cols = [c for c in cols_to_keep if c in self.df2.columns]

            self.df1 = self.df1[df1_cols]
            self.df2 = self.df2[df2_cols]

            self.log.info(f"Filtered to comparing columns: {self.comparing_columns}")
            if self.original_join_cols_source1 or self.original_join_cols_source2:
                self.log.info(f"Kept original join columns for display: source1={list(self.original_join_cols_source1.values())}, source2={list(self.original_join_cols_source2.values())}")

    def validate_duplicates(self) -> bool:
        """
        Check for duplicate rows based on join columns

        Returns:
            True if no duplicates, False otherwise
        """
        self.log.info("Validating for duplicate rows...")

        df1_dups = self.df1.groupby(self.join_columns).size().reset_index(name='records')
        df1_dups = df1_dups[df1_dups.records > 1]

        df2_dups = self.df2.groupby(self.join_columns).size().reset_index(name='records')
        df2_dups = df2_dups[df2_dups.records > 1]

        has_duplicates = False

        if len(df1_dups) > 0:
            self.log.warning(f"Found {len(df1_dups)} duplicate key(s) in {self.data_source1.name}")
            dup_file = os.path.join(self.output_dir, f"duplicates_{self.data_source1.name}.csv")
            df1_dups.to_csv(dup_file, index=False)
            self.log.info(f"Duplicates saved to {dup_file}")
            has_duplicates = True

        if len(df2_dups) > 0:
            self.log.warning(f"Found {len(df2_dups)} duplicate key(s) in {self.data_source2.name}")
            dup_file = os.path.join(self.output_dir, f"duplicates_{self.data_source2.name}.csv")
            df2_dups.to_csv(dup_file, index=False)
            self.log.info(f"Duplicates saved to {dup_file}")
            has_duplicates = True

        if not has_duplicates:
            self.log.info("No duplicates found")

        return not has_duplicates

    def merge_data(self):
        """Merge data from both sources"""
        self.log.info("Merging data from both sources...")

        # Debug: Log columns before merge
        self.log.info(f"df1 columns before merge: {list(self.df1.columns)}")
        self.log.info(f"df2 columns before merge: {list(self.df2.columns)}")

        self.df_merged = pd.merge(
            self.df1,
            self.df2,
            on=self.join_columns,
            how='outer',
            suffixes=[self.suffix1, self.suffix2],
            indicator=True
        )

        # Debug: Log merged columns
        self.log.info(f"Merged columns: {list(self.df_merged.columns)}")

        self.log.info(f"Merged dataframe has {len(self.df_merged)} rows")

        # Rename indicator column
        merge_map = {
            "left_only": f"Only in {self.data_source1.name}",
            "right_only": f"Only in {self.data_source2.name}",
            "both": "Present in Both"
        }
        self.df_merged['_merge'] = self.df_merged['_merge'].map(merge_map)

    def compare_columns(self):
        """
        Compare columns and add comparison result indicators
        """
        self.log.info("Comparing columns...")

        # Get columns to compare (excluding join columns)
        if self.comparing_columns:
            columns_to_compare = [c for c in self.comparing_columns if c not in self.join_columns]
        else:
            # Get all common columns from original dataframes
            columns_to_compare = [c for c in self.df1.columns if c not in self.join_columns]

        # Initialize overall comparison result
        self.df_merged["is_equal"] = True

        # Track column-level comparison results
        for col in columns_to_compare:
            col1 = col + self.suffix1
            col2 = col + self.suffix2

            # Skip if columns don't exist in merged dataframe
            if col1 not in self.df_merged.columns or col2 not in self.df_merged.columns:
                continue

            # Column-specific comparison result
            col_result_name = f"{col}_match"

            # Apply tolerance if specified
            if col in self.tolerance:
                tol = self.tolerance[col]
                self.df_merged[col_result_name] = (
                    (abs(self.df_merged[col1] - self.df_merged[col2]) <= tol) |
                    (pd.isnull(self.df_merged[col1]) & pd.isnull(self.df_merged[col2]))
                )
                self.log.info(f"Applied tolerance {tol} to column '{col}'")
            else:
                # Standard comparison (handles NaN correctly)
                self.df_merged[col_result_name] = (
                    (self.df_merged[col1] == self.df_merged[col2]) |
                    (pd.isnull(self.df_merged[col1]) & pd.isnull(self.df_merged[col2]))
                )

            # Update overall equality
            self.df_merged["is_equal"] = self.df_merged["is_equal"] & self.df_merged[col_result_name]

        # Count matches and differences
        total_rows = len(self.df_merged)
        equal_rows = self.df_merged["is_equal"].sum()
        diff_rows = total_rows - equal_rows

        # Count rows by merge status
        only_in_source1 = len(self.df_merged[self.df_merged['_merge'] == f"Only in {self.data_source1.name}"])
        only_in_source2 = len(self.df_merged[self.df_merged['_merge'] == f"Only in {self.data_source2.name}"])
        in_both = len(self.df_merged[self.df_merged['_merge'] == "Present in Both"])

        self.comparison_summary = {
            "total_rows": total_rows,
            "equal_rows": int(equal_rows),
            "different_rows": int(diff_rows),
            "match_percentage": (equal_rows / total_rows * 100) if total_rows > 0 else 0,
            "only_in_source1": only_in_source1,
            "only_in_source2": only_in_source2,
            "in_both": in_both
        }

        self.log.info(f"Comparison complete: {equal_rows} equal, {diff_rows} different out of {total_rows} total")
        self.log.info(f"Merge status: {in_both} in both, {only_in_source1} only in source1, {only_in_source2} only in source2")

    def generate_csv_report(self, filename: str = "comparison_report.csv"):
        """
        Generate CSV report with merged data and comparison indicators

        Args:
            filename: Output CSV filename
        """
        output_path = os.path.join(self.output_dir, filename)
        self.log.info(f"Generating CSV report: {output_path}")

        self.df_merged.to_csv(output_path, index=False)
        self.log.info(f"CSV report saved to {output_path}")

        return output_path

    def generate_differences_only_csv(self, filename: str = "differences_only.csv"):
        """
        Generate CSV report containing only rows with differences

        Args:
            filename: Output CSV filename
        """
        output_path = os.path.join(self.output_dir, filename)
        self.log.info(f"Generating differences-only CSV: {output_path}")

        df_diff = self.df_merged[self.df_merged["is_equal"] == False]
        df_diff.to_csv(output_path, index=False)

        self.log.info(f"Differences-only CSV saved to {output_path} ({len(df_diff)} rows)")

        return output_path

    def generate_summary_report(self, filename: str = "summary_report.txt"):
        """
        Generate summary text report

        Args:
            filename: Output text filename
        """
        output_path = os.path.join(self.output_dir, filename)
        self.log.info(f"Generating summary report: {output_path}")

        with open(output_path, 'w', encoding='utf-8') as f:
            f.write("=" * 80 + "\n")
            f.write("DATABASE COMPARISON SUMMARY REPORT\n")
            f.write("=" * 80 + "\n\n")

            f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")

            f.write(f"Data Source 1: {self.data_source1.name}\n")
            f.write(f"  Database: {self.data_source1.db_type} - {self.data_source1.database}\n")
            f.write(f"  Rows: {len(self.df1)}\n\n")

            f.write(f"Data Source 2: {self.data_source2.name}\n")
            f.write(f"  Database: {self.data_source2.db_type} - {self.data_source2.database}\n")
            f.write(f"  Rows: {len(self.df2)}\n\n")

            f.write(f"Join Columns: {', '.join(self.join_columns)}\n")
            if self.comparing_columns:
                f.write(f"Comparing Columns: {', '.join(self.comparing_columns)}\n")
            f.write("\n")

            f.write("-" * 80 + "\n")
            f.write("COMPARISON RESULTS\n")
            f.write("-" * 80 + "\n\n")

            f.write(f"Total Rows (after merge): {self.comparison_summary['total_rows']}\n")
            f.write(f"Equal Rows: {self.comparison_summary['equal_rows']}\n")
            f.write(f"Different Rows: {self.comparison_summary['different_rows']}\n")
            f.write(f"Match Percentage: {self.comparison_summary['match_percentage']:.2f}%\n\n")

            f.write("MERGE STATUS:\n")
            f.write(f"  Rows in Both Sources: {self.comparison_summary['in_both']}\n")
            f.write(f"  Only in {self.data_source1.name}: {self.comparison_summary['only_in_source1']}\n")
            f.write(f"  Only in {self.data_source2.name}: {self.comparison_summary['only_in_source2']}\n\n")

            # Overall matching status
            if self.comparison_summary['different_rows'] == 0 and \
               self.comparison_summary['only_in_source1'] == 0 and \
               self.comparison_summary['only_in_source2'] == 0:
                f.write("MATCHING STATUS: PERFECT MATCH [OK]\n")
                f.write("All rows are present in both sources and all values match.\n\n")
            else:
                f.write("MATCHING STATUS: DIFFERENCES FOUND [WARNING]\n")
                if self.comparison_summary['different_rows'] > 0:
                    f.write(f"  - {self.comparison_summary['different_rows']} rows have different values\n")
                if self.comparison_summary['only_in_source1'] > 0:
                    f.write(f"  - {self.comparison_summary['only_in_source1']} rows exist only in {self.data_source1.name}\n")
                if self.comparison_summary['only_in_source2'] > 0:
                    f.write(f"  - {self.comparison_summary['only_in_source2']} rows exist only in {self.data_source2.name}\n")
                f.write("\n")

            # Breakdown by merge indicator
            merge_counts = self.df_merged['_merge'].value_counts()
            f.write("Row Distribution:\n")
            for merge_type, count in merge_counts.items():
                f.write(f"  {merge_type}: {count}\n")

        self.log.info(f"Summary report saved to {output_path}")

        return output_path

    def generate_side_by_side_excel(self, filename: str = "side_by_side_comparison.xlsx"):
        """
        Generate side-by-side Excel report with source1 on left, source2 on right
        Highlights cells where values differ in yellow background

        Args:
            filename: Output Excel filename
        """
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils.dataframe import dataframe_to_rows
        except ImportError:
            self.log.error("openpyxl package is required for Excel export. Install with: pip install openpyxl")
            return None

        output_path = os.path.join(self.output_dir, filename)
        self.log.info(f"Generating side-by-side Excel report: {output_path}")

        # Create a new workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Side by Side Comparison"

        # Define styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        source_name_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")  # Darker blue for source names
        source_name_font = Font(bold=True, color="FFFFFF", size=12)
        diff_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Yellow for differences
        match_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green for matches
        missing_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light red for missing
        key_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # White for keys
        divider_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Gray for divider

        border_thin = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Get columns to display
        comparing_cols = self.comparing_columns if self.comparing_columns else [
            col for col in self.df1.columns if col not in self.join_columns
        ]

        # Calculate column counts for merging cells in row 1 (source names)
        if not self.show_join_columns_both_sides:
            # Keys shown once
            key_cols_count = len(self.join_columns)
            source1_cols_count = len(comparing_cols)
            source2_cols_count = len(comparing_cols)
        else:
            # Keys on both sides
            key_cols_count = 0
            if self.show_transformed_columns:
                # Count original + transformed columns
                source1_extra = sum(1 for jc in self.join_columns
                                   if self.original_join_cols_source1.get(jc, jc) != jc
                                   and self.original_join_cols_source1.get(jc, jc) in self.df1.columns)
                source2_extra = sum(1 for jc in self.join_columns
                                   if self.original_join_cols_source2.get(jc, jc) != jc
                                   and self.original_join_cols_source2.get(jc, jc) in self.df2.columns)
                source1_cols_count = len(self.join_columns) + source1_extra + len(comparing_cols)
                source2_cols_count = len(self.join_columns) + source2_extra + len(comparing_cols)
            else:
                source1_cols_count = len(self.join_columns) + len(comparing_cols)
                source2_cols_count = len(self.join_columns) + len(comparing_cols)

        # Row 1: Data source names
        current_col = 1

        # Keys section (if shown once)
        if not self.show_join_columns_both_sides and key_cols_count > 0:
            # Merge cells for keys section
            if key_cols_count > 1:
                ws.merge_cells(start_row=1, start_column=current_col, end_row=1, end_column=current_col + key_cols_count - 1)
            cell = ws.cell(row=1, column=current_col, value="Keys")
            cell.fill = source_name_fill
            cell.font = source_name_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border_thin
            current_col += key_cols_count

        # Source 1 section
        source1_start_col = current_col
        if source1_cols_count > 1:
            ws.merge_cells(start_row=1, start_column=current_col, end_row=1, end_column=current_col + source1_cols_count - 1)
        cell = ws.cell(row=1, column=current_col, value=self.data_source1.name.upper())
        cell.fill = source_name_fill
        cell.font = source_name_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border_thin
        current_col += source1_cols_count

        # Divider column
        cell = ws.cell(row=1, column=current_col, value="||")
        cell.fill = divider_fill
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border_thin
        divider_col = current_col
        current_col += 1

        # Source 2 section
        source2_start_col = current_col
        if source2_cols_count > 1:
            ws.merge_cells(start_row=1, start_column=current_col, end_row=1, end_column=current_col + source2_cols_count - 1)
        cell = ws.cell(row=1, column=current_col, value=self.data_source2.name.upper())
        cell.fill = source_name_fill
        cell.font = source_name_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border_thin
        current_col += source2_cols_count

        # Match Status section
        cell = ws.cell(row=1, column=current_col, value="Status")
        cell.fill = source_name_fill
        cell.font = source_name_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border_thin

        # Row 2: Column headers
        current_col = 1

        # Helper function to get original column name
        def get_original_col_name_source1(mapped_col):
            """Get original column name from source1 before mapping"""
            # Reverse lookup in column_mapping
            for orig, mapped in self.column_mapping.items():
                if mapped == mapped_col:
                    return orig
            return mapped_col  # Return as-is if not in mapping

        # Build lists of columns to show for each source
        if not self.show_join_columns_both_sides:
            # Keys shown once - just use join column names
            keys_to_show_once = self.join_columns
            source1_cols_to_show = comparing_cols
            source2_cols_to_show = comparing_cols
        else:
            # Keys on both sides
            keys_to_show_once = []
            if self.show_transformed_columns:
                # Show both original and transformed columns
                source1_cols_to_show = []
                source2_cols_to_show = []

                # Add join columns (original + transformed)
                for join_col in self.join_columns:
                    orig1 = self.original_join_cols_source1.get(join_col, join_col)
                    orig2 = self.original_join_cols_source2.get(join_col, join_col)

                    # Source1: add original column (if different) + transformed
                    if orig1 != join_col and orig1 in self.df1.columns:
                        source1_cols_to_show.append(('orig', orig1, orig1))  # (type, display_name, data_col)
                    source1_cols_to_show.append(('transformed', join_col, join_col))

                    # Source2: add original column (if different) + transformed
                    if orig2 != join_col and orig2 in self.df2.columns:
                        source2_cols_to_show.append(('orig', orig2, orig2))
                    source2_cols_to_show.append(('transformed', join_col, join_col))

                # Add comparing columns
                for col in comparing_cols:
                    orig1 = get_original_col_name_source1(col)
                    source1_cols_to_show.append(('comparing', orig1, col))
                    source2_cols_to_show.append(('comparing', col, col))
            else:
                # Show only original columns
                source1_cols_to_show = [(None, self.original_join_cols_source1.get(jc, jc), jc) for jc in self.join_columns]
                source2_cols_to_show = [(None, self.original_join_cols_source2.get(jc, jc), jc) for jc in self.join_columns]
                source1_cols_to_show += [(None, get_original_col_name_source1(col), col) for col in comparing_cols]
                source2_cols_to_show += [(None, col, col) for col in comparing_cols]

        # Write keys shown once (if applicable)
        if not self.show_join_columns_both_sides:
            for join_col in keys_to_show_once:
                cell = ws.cell(row=2, column=current_col, value=join_col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border_thin
                current_col += 1

        # Write Source 1 column headers
        if not self.show_join_columns_both_sides or not self.show_transformed_columns:
            # Simple mode: one column per field
            for col_info in source1_cols_to_show:
                if isinstance(col_info, tuple):
                    _, display_name, _ = col_info
                else:
                    display_name = get_original_col_name_source1(col_info)

                cell = ws.cell(row=2, column=current_col, value=display_name)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border_thin
                current_col += 1
        else:
            # Advanced mode: show original + transformed
            for col_type, display_name, _ in source1_cols_to_show:
                cell = ws.cell(row=2, column=current_col, value=display_name)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border_thin
                current_col += 1

        # Add divider column
        cell = ws.cell(row=2, column=divider_col, value="||")
        cell.fill = divider_fill
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border_thin
        current_col = divider_col + 1

        # Write Source 2 column headers
        if not self.show_join_columns_both_sides or not self.show_transformed_columns:
            # Simple mode: one column per field
            for col_info in source2_cols_to_show:
                if isinstance(col_info, tuple):
                    _, display_name, _ = col_info
                else:
                    display_name = col_info

                cell = ws.cell(row=2, column=current_col, value=display_name)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border_thin
                current_col += 1
        else:
            # Advanced mode: show original + transformed
            for col_type, display_name, _ in source2_cols_to_show:
                cell = ws.cell(row=2, column=current_col, value=display_name)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border_thin
                current_col += 1

        # Add status column
        cell = ws.cell(row=2, column=current_col, value="Match Status")
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border_thin

        # Write data rows (starting from row 3 now)
        row_num = 3
        for idx, row in self.df_merged.iterrows():
            current_col = 1

            # Determine row status
            merge_status = row.get('_merge', 'both')
            is_equal = row.get('is_equal', True)

            # Write join columns once if not showing on both sides
            if not self.show_join_columns_both_sides:
                for join_col in keys_to_show_once:
                    # Get value directly from merged dataframe (join columns don't have suffix)
                    cell = ws.cell(row=row_num, column=current_col, value=row.get(join_col, ''))
                    cell.border = border_thin
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    cell.fill = key_fill  # White background for keys
                    current_col += 1

            # Write source 1 columns
            if not self.show_join_columns_both_sides or not self.show_transformed_columns:
                # Simple mode
                for col_info in source1_cols_to_show:
                    if isinstance(col_info, tuple):
                        _, _, data_col = col_info
                    else:
                        data_col = col_info

                    # Get value
                    if data_col in self.join_columns:
                        value = row.get(data_col, '')  # Join columns don't have suffix
                        is_join_col = True
                    else:
                        col_source1 = f"{data_col}{self.suffix1}"
                        value = row.get(col_source1, '')
                        is_join_col = False

                    cell = ws.cell(row=row_num, column=current_col, value=value)
                    cell.border = border_thin
                    cell.alignment = Alignment(horizontal='left', vertical='center')

                    # Apply background color
                    if is_join_col:
                        cell.fill = key_fill
                    elif merge_status == 'left_only':
                        cell.fill = missing_fill
                    elif merge_status == 'right_only':
                        cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                    else:
                        # Check if column differs
                        col_source2 = f"{data_col}{self.suffix2}"
                        val1 = row.get(f"{data_col}{self.suffix1}")
                        val2 = row.get(col_source2)

                        if pd.isna(val1) and pd.isna(val2):
                            cell.fill = match_fill
                        elif pd.isna(val1) or pd.isna(val2):
                            cell.fill = diff_fill
                        elif val1 != val2:
                            cell.fill = diff_fill
                        else:
                            cell.fill = match_fill

                    current_col += 1
            else:
                # Advanced mode: show original + transformed
                for col_type, _, data_col in source1_cols_to_show:
                    if col_type == 'orig':
                        # Original column - get from df1 directly
                        value = row.get(f"{data_col}{self.suffix1}", '')
                        cell = ws.cell(row=row_num, column=current_col, value=value)
                        cell.fill = key_fill  # White for original
                    elif col_type == 'transformed':
                        # Transformed column
                        value = row.get(data_col, '')
                        cell = ws.cell(row=row_num, column=current_col, value=value)
                        cell.fill = key_fill  # White for transformed key
                    else:
                        # Comparing column
                        col_source1 = f"{data_col}{self.suffix1}"
                        value = row.get(col_source1, '')
                        cell = ws.cell(row=row_num, column=current_col, value=value)

                        # Color based on comparison
                        if merge_status == 'left_only':
                            cell.fill = missing_fill
                        elif merge_status == 'right_only':
                            cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                        else:
                            col_source2 = f"{data_col}{self.suffix2}"
                            val1 = row.get(col_source1)
                            val2 = row.get(col_source2)

                            if pd.isna(val1) and pd.isna(val2):
                                cell.fill = match_fill
                            elif pd.isna(val1) or pd.isna(val2):
                                cell.fill = diff_fill
                            elif val1 != val2:
                                cell.fill = diff_fill
                            else:
                                cell.fill = match_fill

                    cell.border = border_thin
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    current_col += 1

            # Write divider column
            cell = ws.cell(row=row_num, column=divider_col, value="||")
            cell.fill = divider_fill
            cell.border = border_thin
            cell.alignment = Alignment(horizontal='center', vertical='center')
            current_col = divider_col + 1

            # Write source 2 columns
            if not self.show_join_columns_both_sides or not self.show_transformed_columns:
                # Simple mode
                for col_info in source2_cols_to_show:
                    if isinstance(col_info, tuple):
                        _, _, data_col = col_info
                    else:
                        data_col = col_info

                    # Get value
                    if data_col in self.join_columns:
                        value = row.get(data_col, '')  # Join columns don't have suffix
                        is_join_col = True
                    else:
                        col_source2 = f"{data_col}{self.suffix2}"
                        value = row.get(col_source2, '')
                        is_join_col = False

                    cell = ws.cell(row=row_num, column=current_col, value=value)
                    cell.border = border_thin
                    cell.alignment = Alignment(horizontal='left', vertical='center')

                    # Apply background color
                    if is_join_col:
                        cell.fill = key_fill
                    elif merge_status == 'right_only':
                        cell.fill = missing_fill
                    elif merge_status == 'left_only':
                        cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                    else:
                        # Check if column differs
                        col_source1 = f"{data_col}{self.suffix1}"
                        val1 = row.get(col_source1)
                        val2 = row.get(f"{data_col}{self.suffix2}")

                        if pd.isna(val1) and pd.isna(val2):
                            cell.fill = match_fill
                        elif pd.isna(val1) or pd.isna(val2):
                            cell.fill = diff_fill
                        elif val1 != val2:
                            cell.fill = diff_fill
                        else:
                            cell.fill = match_fill

                    current_col += 1
            else:
                # Advanced mode: show original + transformed
                for col_type, _, data_col in source2_cols_to_show:
                    if col_type == 'orig':
                        # Original column - may not have suffix if only in source2
                        # Try with suffix first, then without
                        value = row.get(f"{data_col}{self.suffix2}", row.get(data_col, ''))
                        cell = ws.cell(row=row_num, column=current_col, value=value)
                        cell.fill = key_fill  # White for original
                    elif col_type == 'transformed':
                        # Transformed column
                        value = row.get(data_col, '')
                        cell = ws.cell(row=row_num, column=current_col, value=value)
                        cell.fill = key_fill  # White for transformed key
                    else:
                        # Comparing column
                        col_source2 = f"{data_col}{self.suffix2}"
                        value = row.get(col_source2, '')
                        cell = ws.cell(row=row_num, column=current_col, value=value)

                        # Color based on comparison
                        if merge_status == 'right_only':
                            cell.fill = missing_fill
                        elif merge_status == 'left_only':
                            cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                        else:
                            col_source1 = f"{data_col}{self.suffix1}"
                            val1 = row.get(col_source1)
                            val2 = row.get(col_source2)

                            if pd.isna(val1) and pd.isna(val2):
                                cell.fill = match_fill
                            elif pd.isna(val1) or pd.isna(val2):
                                cell.fill = diff_fill
                            elif val1 != val2:
                                cell.fill = diff_fill
                            else:
                                cell.fill = match_fill

                    cell.border = border_thin
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    current_col += 1

            # Write match status
            if merge_status == 'left_only':
                status = "Only in Source1"
                status_fill = missing_fill
            elif merge_status == 'right_only':
                status = "Only in Source2"
                status_fill = missing_fill
            elif is_equal:
                status = "Match"
                status_fill = match_fill
            else:
                status = "Difference"
                status_fill = diff_fill

            cell = ws.cell(row=row_num, column=current_col, value=status)
            cell.fill = status_fill
            cell.border = border_thin
            cell.alignment = Alignment(horizontal='center', vertical='center')

            row_num += 1

        # Auto-adjust column widths
        for col_idx, column in enumerate(ws.columns, start=1):
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(col_idx)
            for cell in column:
                # Skip merged cells
                if isinstance(cell, openpyxl.cell.cell.MergedCell):
                    continue
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50) if max_length > 0 else 10
            ws.column_dimensions[column_letter].width = adjusted_width

        # Set divider column to narrow width
        ws.column_dimensions[openpyxl.utils.get_column_letter(divider_col)].width = 3

        # Freeze first two rows (source names and column headers)
        ws.freeze_panes = "A3"

        # Save workbook
        wb.save(output_path)
        self.log.info(f"Side-by-side Excel report saved to {output_path}")

        return output_path
        self.log.info(f"Side-by-side Excel report saved to {output_path}")

        return output_path

        return output_path

    def run_comparison(self,
                      generate_full_csv: bool = True,
                      generate_diff_csv: bool = True,
                      generate_summary: bool = True,
                      generate_side_by_side_excel: bool = False,
                      validate_dups: bool = True):
        """
        Run complete comparison workflow

        Args:
            generate_full_csv: Generate full comparison CSV
            generate_diff_csv: Generate differences-only CSV
            generate_summary: Generate summary report
            generate_side_by_side_excel: Generate side-by-side Excel with highlighting
            validate_dups: Validate for duplicate rows

        Returns:
            Dictionary with comparison summary
        """
        start_time = datetime.now()
        self.log.info("=" * 80)
        self.log.info("Starting database comparison")
        self.log.info("=" * 80)

        try:
            # Fetch data
            self.fetch_data()

            # Validate duplicates
            if validate_dups:
                self.validate_duplicates()

            # Merge data
            self.merge_data()

            # Compare columns
            self.compare_columns()

            # Generate reports
            if generate_full_csv:
                self.generate_csv_report()

            if generate_diff_csv:
                self.generate_differences_only_csv()

            if generate_summary:
                self.generate_summary_report()

            if generate_side_by_side_excel:
                self.generate_side_by_side_excel()

            # Calculate elapsed time
            elapsed = (datetime.now() - start_time).total_seconds()
            self.log.info(f"Comparison completed in {elapsed:.2f} seconds")

            # Memory usage
            current, peak = tracemalloc.get_traced_memory()
            self.log.info(f"Memory usage: current {current/1024/1024:.2f}MB, peak {peak/1024/1024:.2f}MB")

            return self.comparison_summary

        except Exception as e:
            self.log.exception("Comparison failed with error:")
            raise

        finally:
            # Close connections
            self.data_source1.close()
            self.data_source2.close()


def create_data_source_from_dict(config: Dict) -> DataSource:
    """
    Helper function to create DataSource from dictionary configuration

    Args:
        config: Dictionary with data source configuration

    Returns:
        DataSource object
    """
    return DataSource(
        name=config.get('name', 'datasource'),
        db_type=config['db_type'],
        host=config.get('host', 'localhost'),
        port=config.get('port', 3306),
        database=config['database'],
        username=config['username'],
        password=config['password'],
        sql_query=config['sql_query'],
        **config.get('connection_params', {})
    )


if __name__ == "__main__":
    # Example usage
    print(f"DB Compare version: {version}")
    print("This is a library module. Import DBCompare and DataSource classes to use.")
    print("\nExample:")
    print("""
    from db_compare import DBCompare, DataSource
    
    # Configure data sources
    source1 = DataSource(
        name="production",
        db_type="mysql",
        host="localhost",
        port=3306,
        database="prod_db",
        username="user",
        password="pass",
        sql_query="SELECT id, name, value FROM table1"
    )
    
    source2 = DataSource(
        name="staging",
        db_type="mysql",
        host="localhost",
        port=3306,
        database="stage_db",
        username="user",
        password="pass",
        sql_query="SELECT id, name, value FROM table1"
    )
    
    # Run comparison
    comparator = DBCompare(
        data_source1=source1,
        data_source2=source2,
        join_columns=["id"],
        comparing_columns=["name", "value"]
    )
    
    results = comparator.run_comparison()
    print(results)
    """)

